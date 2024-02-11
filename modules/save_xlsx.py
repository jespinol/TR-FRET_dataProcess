import os
import time

import openpyxl.chart.label
import openpyxl.utils.cell
import pandas as pd
from openpyxl.styles import Font

from modules.plotting import *


def output_results(data_info, signal_corrected, signal_normalized, fit_data):
    path = data_info[PATH]
    workbook = create_workbook(path)
    writer = pd.ExcelWriter(workbook, engine="openpyxl", mode="a", if_sheet_exists="overlay")

    # add normalized data to the output file
    row = 0
    write_signal_data(signal_normalized, writer, WS_NAME, row)

    # add unnormalized data to output file
    row = data_info[NUM_DATAPOINTS] + 5
    write_signal_data(signal_corrected, writer, WS_NAME, row)

    # add theoretical data based on models
    col = 17
    for model in MODELS:
        fitted_curve_datapoints = calculate_fitted_curve_datapoints(signal_normalized, fit_data, model)
        df_fitted_data = pd.DataFrame(data=fitted_curve_datapoints)
        df_fitted_data.to_excel(writer, sheet_name=WS_NAME, index=False, startcol=col)
        col += 3

    # add helper series for chart x-axis labels
    df_helper_data = pd.DataFrame(
        data={HELPER_X: [1e-1, 1e0, 1e1, 1e2, 1e3, 1e4, 1e5], HELPER_Y: [-0.1, -0.1, -0.1, -0.1, -0.1, -0.1, -0.1],
              HELPER_LABEL: [f"{' ' * 10}-1", f"{' ' * 9}0", f"{' ' * 9}1", f"{' ' * 9}2", f"{' ' * 9}3", f"{' ' * 9}4",
                             f"{' ' * 9}5"]})
    df_helper_data.to_excel(writer, sheet_name=WS_NAME, index=False, startcol=col)

    # add plots of log(concentration) vs. normalized signal including the fitted curve according to each model
    plot_worksheet = writer.sheets[WS_NAME]
    row = 2
    chart_col = chr(DEFAULT_COL_NUM + data_info[NUM_REPEATS] + ord("A"))
    fit_col = DEFAULT_COL_NUM + data_info[NUM_REPEATS]
    for model in MODELS:
        chart = create_chart(plot_worksheet, model)
        plot_worksheet.add_chart(chart, f"{chart_col}{row}")
        row += CHART_ROW_HEIGHT + 3

        # add curve fitting data
        title_cell = plot_worksheet[f"{chr(fit_col + ord('A'))}{row}"]
        title_cell.value = model
        title_cell.font = openpyxl.styles.Font(bold=True)
        df_fit = create_fit_DataFrame(fit_data[model])
        df_fit.to_excel(writer, sheet_name=WS_NAME, index=False, header=True, startrow=row, startcol=fit_col)

        row += 7

    writer.close()

    return


def create_workbook(path):
    # add a suffix with the current time so the file will always be unique
    filename_suffix = time.strftime("%Y%m%d-%H%M%S")

    # if a single input file was provided, output will have the same filename with xlsx extension
    # if a directory was provided as input, the output name will have the name of the directory
    if path.endswith(".csv"):
        root, extension = os.path.splitext(path)
        filename = f"{root}_{filename_suffix}.xlsx"
    else:
        parent_dir = os.path.basename(path)
        filename = f"{path}/{parent_dir}_{filename_suffix}.xlsx"

    # open a new workbook and rename the default worksheet name
    workbook = openpyxl.Workbook()
    workbook["Sheet"].title = WS_NAME

    # increase the width of a number of columns
    resize_columns(workbook)

    workbook.save(filename)

    # returns the path to the output file so that ExcelWriter and other writing functions can find it
    return filename


def resize_columns(workbook):
    for worksheet in workbook.sheetnames:
        sheet = workbook[worksheet]
        for col_idx in range(1, 40):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 17

    return


def write_signal_data(data, writer, worksheet, row=0, col=0):
    df = create_signal_DataFrame(data)
    df.to_excel(writer, sheet_name=worksheet, index=False, startrow=row, startcol=col)


def create_signal_DataFrame(signal_data):
    data = {}
    for column in SIGNAL_DATAFRAME_FORMAT:
        if column == SIGNAL_VALUES:
            for repeat in signal_data[column]:
                label = f"Repeat {repeat}"
                data[label] = signal_data[column][repeat]
        elif column == STATS:
            for stat in signal_data[column]:
                data[stat] = signal_data[column][stat]
        else:
            data[column] = signal_data[column]

    data = {**{CONC: signal_data[CONC]}, **data}
    df = pd.DataFrame(data)
    return df


def create_fit_DataFrame(fit_data):
    data = {PARAMETER: [], VALUE: [], CONF_INT_LOWER: [], CONF_INT_UPPER: [], STD_DEV: [], STD_ERR: []}
    for param, value in fit_data.items():
        data[PARAMETER].append(param)
        data[VALUE].append(value[PARAMETER])
        data[CONF_INT_LOWER].append(value[CONF_INT_LOWER])
        data[CONF_INT_UPPER].append(value[CONF_INT_UPPER])
        data[STD_DEV].append(value[STD_DEV])
        data[STD_ERR].append(value[STD_ERR])

    df = pd.DataFrame(data)
    return df


def calculate_fitted_curve_datapoints(signal_data, fit_data, model):
    max_concentration = signal_data[CONC][0] * 1.2
    min_concentration = signal_data[CONC][-1] * 0.8
    kd = fit_data[model][KD][PARAMETER]
    model_equation = MODEL_EQUATIONS[model]

    fit_x_values = []
    fit_y_values = []

    current_x = max_concentration
    while current_x > min_concentration:
        fit_x_values.append(current_x)
        if model == COOPERATIVE_MODEL:
            nH = fit_data[model][NH][PARAMETER]
            y_val = model_equation(current_x, kd, nH)
        else:
            y_val = model_equation(current_x, kd)
        fit_y_values.append(y_val)
        if np.any(np.isnan(y_val)):
            fit_x_values = fit_y_values = []
            break
        current_x *= 0.9

    x_values_label = f"x_{model}"
    y_values_label = f"y_{model}"

    return {x_values_label: fit_x_values, y_values_label: fit_y_values}
